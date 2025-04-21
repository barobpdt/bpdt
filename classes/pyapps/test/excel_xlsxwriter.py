import openpyxl
import xlsxwriter

# Load the workbook and sheet
workbook = openpyxl.load_workbook("C:/Users/xyz/Downloads/DEMO_I.xlsx")
sheet = workbook["DEMO_I"]

# Create a new Excel workbook and worksheet using XlsxWriter
output_file = "C:/Users/xyz/Downloads/DEMO_I_protected.xlsx"
workbook_xlsx = xlsxwriter.Workbook(output_file)
worksheet_xlsx = workbook_xlsx.add_worksheet()

# Copy data from existing worksheet to new worksheet
for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    for col_num, value in enumerate(row, start=1):
        worksheet_xlsx.write(row_num - 1, col_num - 1, value)

# Set unlocked format for all cells
unlocked_format = workbook_xlsx.add_format({'locked': False})
worksheet_xlsx.protect("1234", options={'format_columns': True, 'autofilter': True, 'sort': True,
                   'select_locked_cells':False,'select_unlocked_cells':False,'delete_columns':False,
                   'insert_columns':False,'delete_rows':False,'format_cells':False,'format_rows':False})

# Apply autofilter to all columns
num_rows = sheet.max_row
num_cols = sheet.max_column
worksheet_xlsx.autofilter(0, 0, num_rows - 1, num_cols - 1)

# Apply sort to all columns
worksheet_xlsx.unprotect_range('A1:D12', 'MyRange')

# Save the XlsxWriter workbook
workbook_xlsx.close()